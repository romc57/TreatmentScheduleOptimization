import random
from collections import defaultdict
from openpyxl import Workbook
from itertools import cycle

PROFESSIONS = ["nurse", "doctor", "therapist", "psychologist", "care_assistant"]
DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
HOURS = list(range(8, 18))
NAMES_POOL = [
    "Alice", "Bob", "Charlie", "Diana", "Eli", "Fiona", "George", "Hannah", "Ivan", "Julia",
    "Kevin", "Laura", "Mike", "Nina", "Oscar", "Paula", "Quinn", "Rita", "Steve", "Tina",
    "Uma", "Victor", "Wendy", "Xander", "Yara", "Zane", "Abby", "Ben", "Carmen", "Derek",
    "Esther", "Frank", "Gina", "Harold", "Isla", "Jake", "Karen", "Leo", "Mila", "Noah"
]

class Caretaker:
    def __init__(self, name, profession, working_days, working_hours):
        self.name = f"{name} ({profession})"
        self.profession = profession
        self.working_days = working_days
        self.working_hours = working_hours

class Patient:
    def __init__(self, pid):
        self.pid = pid
        self.assignments = []

    def add_assignment(self, day, hour, caretaker_name, profession):
        self.assignments.append((day, hour, caretaker_name, profession))

class Scheduler:
    def __init__(self, num_caretakers=30, num_patients=80):
        self.num_caretakers = num_caretakers
        self.num_patients = num_patients
        self.caretakers = []
        self.patients = []

    def generate_caretakers(self):
        used_names = set()
        prof_cycle = cycle(PROFESSIONS)
        for _ in range(self.num_caretakers):
            while True:
                name = random.choice(NAMES_POOL)
                if name not in used_names:
                    used_names.add(name)
                    break
            profession = next(prof_cycle)
            working_days = random.sample(DAYS, random.randint(3, 6))
            block_length = random.randint(4, 8)
            start_hour = random.randint(8, 17 - block_length + 1)
            working_hours = list(range(start_hour, start_hour + block_length))
            self.caretakers.append(Caretaker(name, profession, working_days, working_hours))

    def create_patient_workbook(self):
        wb = Workbook()
        wb.remove(wb.active)
        for patient in self.patients:
            grid = [["" for _ in DAYS] for _ in HOURS]
            for d, h, cname, prof in patient.assignments:
                r = HOURS.index(h)
                c = DAYS.index(d)
                grid[r][c] = f"{prof} ({cname.split()[0]})"
            ws = wb.create_sheet(title=patient.pid)
            ws.cell(row=1, column=1, value="Hour")
            for col, day in enumerate(DAYS, 2):
                ws.cell(row=1, column=col, value=day)
            for row, hour in enumerate(HOURS, 2):
                ws.cell(row=row, column=1, value=hour)
                for col in range(2, len(DAYS) + 2):
                    ws.cell(row=row, column=col, value=grid[row - 2][col - 2])
        path = "patient_schedule_with_names.xlsx"
        wb.save(path)
        return path

    def create_caretaker_workbook(self):
        wb = Workbook()
        wb.remove(wb.active)
        schedule = defaultdict(lambda: [["" for _ in DAYS] for _ in HOURS])
        for patient in self.patients:
            for d, h, cname, _ in patient.assignments:
                r = HOURS.index(h)
                c = DAYS.index(d)
                schedule[cname][r][c] = patient.pid
        for cname, grid in schedule.items():
            ws = wb.create_sheet(title=cname[:31])
            ws.cell(row=1, column=1, value="Hour")
            for col, day in enumerate(DAYS, 2):
                ws.cell(row=1, column=col, value=day)
            for row, hour in enumerate(HOURS, 2):
                ws.cell(row=row, column=1, value=hour)
                for col in range(2, len(DAYS) + 2):
                    ws.cell(row=row, column=col, value=grid[row - 2][col - 2])
        path = "caretaker_schedule_oop.xlsx"
        wb.save(path)
        return path

    def export_json(self):
        # Caretaker JSON: {caretaker: {day: {hour: patient}}}
        caretaker_json = {}
        for ct in self.caretakers:
            grid = defaultdict(dict)
            for patient in self.patients:
                for d, h, cname, _ in patient.assignments:
                    if cname == ct.name:
                        grid[d][h] = patient.pid
            caretaker_json[ct.name] = {day: hours for day, hours in grid.items()}

        # Patient JSON: {patient: {day: {hour: caretaker}}}
        patient_json = {}
        for patient in self.patients:
            grid = defaultdict(dict)
            for d, h, cname, _ in patient.assignments:
                grid[d][h] = cname
            patient_json[patient.pid] = {day: hours for day, hours in grid.items()}

        import json
        with open("caretaker_schedule.json", "w") as f:
            json.dump(caretaker_json, f, indent=2)
        with open("patient_schedule.json", "w") as f:
            json.dump(patient_json, f, indent=2)
        print("Caretaker JSON written to caretaker_schedule.json")
        print("Patient JSON written to patient_schedule.json")
        return caretaker_json, patient_json

def enforce_consistent_caretakers_per_profession(scheduler):
    for patient in scheduler.patients:
        profession_to_caretaker = {}
        updated_assignments = []
        for d, h, cname, prof in patient.assignments:
            if prof not in profession_to_caretaker:
                profession_to_caretaker[prof] = cname
            updated_assignments.append((d, h, profession_to_caretaker[prof], prof))
        patient.assignments = updated_assignments

class StrictScheduler(Scheduler):
    def assign_patients(self):
        self.patients = []
        patient_id = 1
        caretaker_assignments = defaultdict(lambda: defaultdict(dict))

        for ct in self.caretakers:
            for day in ct.working_days:
                for hour in ct.working_hours:
                    assigned = False
                    attempt = 0
                    while not assigned and attempt < 100:
                        if len(self.patients) < self.num_patients:
                            patient = Patient(f"P{patient_id:03d}")
                            self.patients.append(patient)
                            patient_id += 1
                        else:
                            patient = random.choice(self.patients)

                        same_day_profs = [p for d, _, _, p in patient.assignments if d == day]
                        if ct.profession not in same_day_profs:
                            patient.add_assignment(day, hour, ct.name, ct.profession)
                            caretaker_assignments[ct.name][day][hour] = patient.pid
                            assigned = True
                        attempt += 1

# --- Example Usage ---

if __name__ == "__main__":
    scheduler = StrictScheduler(num_caretakers=30, num_patients=80)
    scheduler.generate_caretakers()
    scheduler.assign_patients()
    enforce_consistent_caretakers_per_profession(scheduler)

    # Export schedules to Excel
    patient_excel = scheduler.create_patient_workbook()
    caretaker_excel = scheduler.create_caretaker_workbook()

    # Export schedules to JSON
    scheduler.export_json()

    print(f"Patient schedule saved to: {patient_excel}")
    print(f"Caretaker schedule saved to: {caretaker_excel}")
